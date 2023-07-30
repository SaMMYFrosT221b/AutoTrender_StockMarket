import openpyxl
import pandas as pd
import requests

wb = openpyxl.load_workbook("./Option.xlsx")

total_sheets = wb.sheetnames
sh1 = wb["Sheet1"]


URL = "https://www.nseindia.com/api/option-chain-indices?symbol=NIFTY"

header = {
    "Accept-Encoding": "gzip, deflate, br",
    "Accept-Language": "en-US,en;q=0.9",
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/114.0.0.0 Safari/537.36",
}


session = requests.session()

data = session.get(url=URL, headers=header).json()["records"]["data"]

df = pd.DataFrame(data)

count = 0
list = [
    "openInterest",
    "changeinOpenInterest",
    "totalTradedVolume",
    "lastPrice",
    "change",
    "strikePrice",
]
CE_data = []
PE_data = []
for i in range(len(df)):
    if df["expiryDate"][i] == "03-Aug-2023":
        CE_data.append(df["CE"][i])
        PE_data.append(df["PE"][i])


openInterestCE = []
changeinOpenInterestCE = []
totalTradedVolumeCE = []
# lastPriceCE = []
# changeCE = []
strikePrice = []
openInterestPE = []
changeinOpenInterestPE = []
totalTradedVolumePE = []
# lastPricePE = []
# changePE = []


# print(CE_data)

for i in range(len(CE_data)):
    openInterestCE.append(CE_data[i]["openInterest"])
    changeinOpenInterestCE.append(CE_data[i]["changeinOpenInterest"])
    totalTradedVolumeCE.append(CE_data[i]["totalTradedVolume"])
    # lastPriceCE.append(CE_data[i]["lastPrice"])
    # changeCE.append(CE_data[i]["change"])

    strikePrice.append(CE_data[i]["strikePrice"])

    openInterestPE.append(PE_data[i]["openInterest"])
    changeinOpenInterestPE.append(PE_data[i]["changeinOpenInterest"])
    totalTradedVolumePE.append(PE_data[i]["totalTradedVolume"])
    # lastPricePE.append(PE_data[i]["lastPrice"])
    # changePE.append(PE_data[i]["change"])


for i in range(2, len(CE_data)):
    sh1.cell(i, 1).value = openInterestCE[i - 2]
    sh1.cell(i, 2).value = changeinOpenInterestCE[i - 2]
    sh1.cell(i, 3).value = totalTradedVolumeCE[i - 2]
    sh1.cell(i, 4).value = strikePrice[i - 2]
    sh1.cell(i, 5).value = totalTradedVolumePE[i - 2]
    sh1.cell(i, 6).value = changeinOpenInterestPE[i - 2]
    sh1.cell(i, 7).value = openInterestPE[i - 2]


wb.save("./Option.xlsx")
wb.close()
